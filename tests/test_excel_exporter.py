import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from web_ders_programi.excel_exporter import export_schedule
from web_ders_programi.models import Course
from web_ders_programi.scheduler import build_department_schedule


class ExcelExporterTests(unittest.TestCase):
    def test_excel_dosyasi_olusturulur(self):
        courses = [
            Course("S018", "Alan Seçmeli Ders I", 2, "Yazılım Mühendisliği", 6, "Yönetici", "D201")
        ]
        entries = build_department_schedule(courses, "Yazılım Mühendisliği")

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "program.xlsx"
            export_schedule(entries, "Yazılım Mühendisliği", path)
            self.assertTrue(path.exists())

            wb = load_workbook(path)
            ws = wb.active
            values = [cell.value for row in ws.iter_rows() for cell in row]
            joined = "\n".join(str(value) for value in values if value)
            self.assertIn("S018 - Alan Seçmeli Ders I", joined)
            self.assertNotIn("Öğretim Üyesi: Yönetici", joined)


if __name__ == "__main__":
    unittest.main()
