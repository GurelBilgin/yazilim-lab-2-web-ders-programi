from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import ScheduleEntry
from .scheduler import DAYS, HOURS


def create_empty_workbook(department: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ders Programı"

    ws.append(["Bölüm", "", department, "", "", ""])
    ws.append(["Gün", "Saat", "1. Sınıf", "2. Sınıf", "3. Sınıf", "4. Sınıf"])

    for day in DAYS:
        first = True
        for hour in HOURS:
            ws.append([day if first else "", hour, "", "", "", ""])
            first = False

    _style_sheet(ws)
    return wb


def export_schedule(entries: list[ScheduleEntry], department: str, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = create_empty_workbook(department)
    ws = wb.active

    slot_to_row = {}
    row = 3
    for day in DAYS:
        for hour in HOURS:
            slot_to_row[(day, hour)] = row
            row += 1

    for entry in entries:
        target_row = slot_to_row[(entry.day, entry.hour)]
        target_col = 2 + entry.class_column
        cell = ws.cell(row=target_row, column=target_col)
        cell.value = entry.cell_text
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[target_row].height = max(ws.row_dimensions[target_row].height or 48, 58)

    wb.save(output_path)
    return output_path


def _style_sheet(ws) -> None:
    widths = [18, 18, 34, 34, 34, 34]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    sub_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="FFFFFF")
    sub_font = Font(bold=True, color="000000")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for cell in ws[2]:
        cell.fill = sub_fill
        cell.font = sub_font

    for idx in range(3, ws.max_row + 1):
        ws.row_dimensions[idx].height = 42

    ws.freeze_panes = "A3"
